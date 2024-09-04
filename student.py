from flask import Flask, render_template, request, jsonify  # type: ignore
import openpyxl  # type: ignore
import os

app = Flask(__name__)

# Path to the Excel file
EXCEL_FILE = 'students.xlsx'

# Create the Excel file if it doesn't exist
if not os.path.exists(EXCEL_FILE):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Attendance'
    sheet.append(['ID', 'Name', 'Attendance'])
    workbook.save(EXCEL_FILE)

def load_students():
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet = workbook['Attendance']
    students = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        students.append({"id": row[0], "name": row[1], "attendance": row[2]})
    return students

def save_student(student):
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet = workbook['Attendance']
    sheet.append([student['id'], student['name'], student['attendance']])
    workbook.save(EXCEL_FILE)

def update_student(student_id, updated_student):
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet = workbook['Attendance']
    for row in sheet.iter_rows(min_row=2):
        if row[0].value == student_id:
            row[1].value = updated_student["name"]
            row[2].value = updated_student["attendance"]
            workbook.save(EXCEL_FILE)
            return True
    return False

def delete_student(student_id):
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet = workbook['Attendance']
    for row in sheet.iter_rows(min_row=2):
        if row[0].value == student_id:
            sheet.delete_rows(row[0].row)
            workbook.save(EXCEL_FILE)
            return True
    return False

@app.route('/')
def index():
    return render_template("student.html")

@app.route('/students', methods=['GET'])
def get_students():
    students = load_students()
    return jsonify(students)

@app.route('/students', methods=['POST'])
def add_student():
    new_student = request.json
    save_student(new_student)
    return jsonify(new_student), 201

@app.route('/students/<int:student_id>', methods=['PUT'])
def update_student_info(student_id):
    updated_student = request.json
    if update_student(student_id, updated_student):
        return jsonify(updated_student)
    else:
        return jsonify({"error": "Student not found"}), 404

@app.route('/students/<int:student_id>', methods=['DELETE'])
def remove_student(student_id):
    if delete_student(student_id):
        return '', 204
    else:
        return jsonify({"error": "Student not found"}), 404

if __name__ == '__main__':
    app.run(debug=True)
